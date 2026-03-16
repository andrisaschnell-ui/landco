"""
importers/import_bdo.py
Read BDO Bank Control .xlsx files (Summary, Creditors, Invoices, BIM sheets).
"""
import os, glob
import openpyxl
from datetime import datetime
from rich.console import Console
import sys; sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..'))
from config import FOLDERS
from db import get_connection

console = Console()
BDO_DIR = FOLDERS["bdo"]

BANK_ACCOUNT_MAP = {
    "BDO":       1,
    "BIM USD":   2,
    "BIM MTN":   3,
    "PETTY":     4,
}


def _safe(val) -> float:
    try:
        if val is None or str(val).strip() in ("#REF!", "#N/A", "#VALUE!", ""):
            return 0.0
        return float(str(val).replace(",", "").strip())
    except Exception:
        return 0.0


def _parse_month_from_filename(fname: str) -> int:
    base = os.path.basename(fname)
    try:
        return int(base.split(" ")[0])
    except Exception:
        return 0


def import_bdo_file(filepath: str):
    month = _parse_month_from_filename(filepath)
    year = 2026
    console.print(f"[cyan]BDO: {os.path.basename(filepath)} (month {month})[/cyan]")

    wb = openpyxl.load_workbook(filepath, data_only=True)
    conn = get_connection()
    try:
        cur = conn.cursor()
        total_loaded = 0

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            sheet_up = sheet_name.strip().upper()

            # Determine bank account
            bank_account_id = None
            currency = "MZN"
            if "BIM" in sheet_up and "USD" in sheet_up:
                bank_account_id = 2
                currency = "USD"
            elif "BIM" in sheet_up and "MTN" in sheet_up:
                bank_account_id = 3
                currency = "MZN"
            elif "PETTY" in sheet_up:
                bank_account_id = 4
                currency = "MZN"
            else:
                # Default: BDO MZN
                bank_account_id = 1

            # Find date-based rows
            for row in ws.iter_rows(min_row=5, values_only=True):
                if not row or row[0] is None:
                    continue
                if not isinstance(row[0], datetime):
                    continue

                tx_date = row[0].date()
                # Try to extract desc, credit, debit, balance from row
                # Columns vary per sheet — use heuristic: find first non-None values
                non_none = [i for i, c in enumerate(row) if c is not None and i > 0]
                if len(non_none) < 2:
                    continue

                desc = ""
                credit = debit = balance = 0.0
                # Simple mapping based on common column patterns
                vals = [_safe(row[i]) if i < len(row) else 0.0 for i in range(1, min(12, len(row)))]
                if len(vals) >= 4:
                    credit  = vals[1] if vals[1] else 0.0
                    debit   = vals[2] if vals[2] else 0.0
                    balance = vals[3] if vals[3] else 0.0

                for i in non_none:
                    cell = row[i]
                    if isinstance(cell, str) and len(cell) > 2:
                        desc = cell.strip()
                        break

                if credit == 0 and debit == 0:
                    continue

                cur.execute("""
                    INSERT INTO bank_transactions
                      (bank_account_id, tx_date, month, year,
                       description, credit, debit, balance,
                       source_file)
                    VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s)
                """, (bank_account_id, tx_date, month, year,
                      desc, credit, debit, balance,
                      os.path.basename(filepath)))
                total_loaded += 1

        conn.commit()
        console.print(f"  [green]✓ {total_loaded} bank transaction rows loaded[/green]")
    finally:
        conn.close()


def run():
    console.print("[bold blue]═══ BDO Bank Control Importer ═══[/bold blue]")
    files = sorted(glob.glob(os.path.join(BDO_DIR, "*.xlsx")))
    if not files:
        console.print(f"[yellow]No BDO files found in {BDO_DIR}[/yellow]")
        return
    for f in files:
        import_bdo_file(f)


if __name__ == "__main__":
    run()
