"""
importers/import_accounts.py
Read Landco Month-End account files (INCOME, EXPENSES, RECON sheets).
"""
import os, glob, re
import openpyxl
from datetime import datetime
from rich.console import Console
import sys; sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..'))
from config import FOLDERS, month_from_str
from db import get_connection

console = Console()

ACCOUNTS_DIR = FOLDERS["accounts"]

HOUSE_MAP = {"H1": 1, "H2": 2, "H3": 3, "H4": 4}


def _safe(val) -> float:
    try:
        if val is None or str(val).strip() in ("#REF!", "#N/A", "#VALUE!", ""):
            return 0.0
        return float(str(val).replace(",", "").strip())
    except Exception:
        return 0.0


def _parse_month_from_path(path: str) -> int:
    """Extract month from folder name like '01 JAN ACCOUNTS'"""
    parts = path.replace("\\", "/").split("/")
    for part in parts:
        m = re.match(r'^(\d{2})\s', part)
        if m:
            return int(m.group(1))
    return 0


def import_month_end_file(filepath: str):
    month = _parse_month_from_path(filepath)
    year = 2026
    console.print(f"[cyan]Month-end: {os.path.basename(filepath)} (month {month})[/cyan]")

    wb = openpyxl.load_workbook(filepath, data_only=True)

    # ── INCOME sheet ───────────────────────────────────────────
    if "INCOME" in wb.sheetnames:
        ws = wb["INCOME"]
        conn = get_connection()
        try:
            cur = conn.cursor()
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row or row[3] is None:
                    continue
                house_code = str(row[3]).strip().upper()
                prop_id = HOUSE_MAP.get(house_code)
                if not prop_id:
                    continue
                amount_mzn = _safe(row[4])
                amount_usd = _safe(row[5]) if len(row) > 5 else 0.0
                if amount_mzn == 0 and amount_usd == 0:
                    continue
                cur.execute("""
                    INSERT INTO income_transactions
                      (property_id, month, year, description, amount_mzn, amount_usd, source_file)
                    VALUES (%s,%s,%s,%s,%s,%s,%s)
                """, (prop_id, month, year, f"Income {house_code}",
                      amount_mzn, amount_usd or None,
                      os.path.basename(filepath)))
            conn.commit()
            console.print("  [green]✓ Income rows loaded[/green]")
        finally:
            conn.close()

    # ── RECON sheet ────────────────────────────────────────────
    if "RECON" in wb.sheetnames:
        ws = wb["RECON"]
        conn = get_connection()
        try:
            cur = conn.cursor()
            for row in ws.iter_rows(min_row=3, values_only=True):
                if not row or row[1] is None:
                    continue
                house_name = str(row[1]).strip().upper()
                # Map name to property_id
                prop_id = None
                if "LUZ"    in house_name: prop_id = 1
                elif "COCO"  in house_name: prop_id = 2
                elif "AURORA"in house_name: prop_id = 3
                elif "CAJU"  in house_name: prop_id = 4
                if not prop_id:
                    continue

                opening  = _safe(row[2])
                income   = _safe(row[3])
                expenses = _safe(row[4])
                closing  = _safe(row[5])

                cur.execute("""
                    INSERT INTO monthly_recon
                      (property_id, month, year,
                       opening_balance_mzn, total_income_mzn,
                       total_expenses_mzn,  closing_balance_mzn,
                       source_file)
                    VALUES (%s,%s,%s,%s,%s,%s,%s,%s)
                    ON DUPLICATE KEY UPDATE
                      opening_balance_mzn  = VALUES(opening_balance_mzn),
                      total_income_mzn     = VALUES(total_income_mzn),
                      total_expenses_mzn   = VALUES(total_expenses_mzn),
                      closing_balance_mzn  = VALUES(closing_balance_mzn)
                """, (prop_id, month, year, opening, income, expenses, closing,
                      os.path.basename(filepath)))
            conn.commit()
            console.print("  [green]✓ Recon rows loaded[/green]")
        finally:
            conn.close()


def run():
    console.print("[bold blue]═══ Month-End Accounts Importer ═══[/bold blue]")
    pattern = os.path.join(ACCOUNTS_DIR, "**", "*MONTH END*.xlsx")
    files = sorted(glob.glob(pattern, recursive=True))
    if not files:
        console.print(f"[yellow]No month-end files found in {ACCOUNTS_DIR}[/yellow]")
        return
    for f in files:
        import_month_end_file(f)


if __name__ == "__main__":
    run()
