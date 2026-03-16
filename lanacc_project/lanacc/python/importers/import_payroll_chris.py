"""
import_payroll_chris.py - Import payroll data from the Chris spreadsheets
"""
import os
from typing import List, Tuple

import openpyxl
from dotenv import load_dotenv

from db import execute_many, execute_one, fetch_all

BASE_DIR = os.path.abspath(os.path.join(os.path.dirname(__file__), "..", ".."))
load_dotenv(os.path.join(BASE_DIR, ".env"))

PAYROLL_SOURCE_DIR = os.getenv(
    "PAYROLL_SOURCE_DIR",
    r"C:\Users\Andrisa\Dropbox\Andrisa\Landco\Chris new program",
)

SHEET_NAME = "Folha de salarios"

FILES = {
    1: ("01 BDO Bank Control 2026 Chris.xlsx", "Jan26"),
    2: ("02 BDO Bank Control 2026 Chris.xlsx", "Feb26"),
}


def _find_header_row(ws) -> int:
    """Use the fixed header row requested by Andrisa."""
    return 6


def _build_headers(ws, header_row: int) -> List[str]:
    """Build display headers using only the fixed header row."""
    row = [c.value for c in ws[header_row]]
    headers = []
    for v in row:
        if v is not None and str(v).strip() != "":
            headers.append(str(v).strip())
        else:
            headers.append("")
    last = 0
    for i, h in enumerate(headers, start=1):
        if str(h).strip() != "":
            last = i
    if last == 0:
        return []
    return headers[:last]


def _read_rows(ws, start_row: int, col_count: int) -> List[List]:
    rows = []
    empty_streak = 0
    for r in range(start_row, ws.max_row + 1):
        row = [ws.cell(row=r, column=c).value for c in range(1, col_count + 1)]
        if all(v is None or str(v).strip() == "" for v in row):
            empty_streak += 1
            if empty_streak >= 5:
                break
            continue
        empty_streak = 0
        rows.append(row)
    return rows


def _clean_rows(rows, no_idx, name_idx):
    """Remove section/header rows and round numeric amounts to 0 decimals."""
    banned_tokens = [
        "data :",
        "declaracao",
        "declaro que nesta data",
        "folha, aos trabalhadores nela mencionados",
        "que comigo, pagador, vao assinar",
        "riscar o que nao interessa",
        "assinalar com m os trabalhadores de menor idade",
        "nao preencher quando se trata de trabalhadores ao dia",
        "descontos permitidos por lei",
        "7% inss",
        "4% inss",
        "315177",
        "360332",
        "24253",
        "13859",
        "330886",
        "5668",
        "84",
        "9919",
        "346473",
        "7000",
        "10438",
        "10394",
        "3465",
        "31296",
        "( a )",
        "( b )",
        "( c )",
        "( d )",
    ]
    cleaned = []
    for row in rows:
        # filter rows that contain banned tokens
        row_text = " ".join([str(v) for v in row if v is not None]).lower()
        if any(tok in row_text for tok in banned_tokens):
            continue
        if no_idx >= 0:
            no_val = "" if row[no_idx] is None else str(row[no_idx]).strip()
            if no_val.lower() == "no":
                continue
            if no_val == "" or not no_val.isdigit():
                continue

        new_row = []
        for idx, v in enumerate(row):
            if idx == name_idx and isinstance(v, str):
                new_row.append(" ".join([w.capitalize() for w in v.strip().lower().split()]))
            elif isinstance(v, (int, float)):
                new_row.append(int(round(v, 0)))
            else:
                new_row.append(v)
        cleaned.append(new_row)
    return cleaned


def _sanitize_headers(headers: List[str]) -> Tuple[List[str], int]:
    """
    Return sanitized DB column names and the index of the NO column.
    """
    col_names = []
    no_idx = -1
    for i, h in enumerate(headers):
        col = f"col_{i+1}"
        col_names.append(col)
        if str(h).strip().lower() == "no":
            no_idx = i
    return col_names, no_idx


def _ensure_meta_table():
    sql = """
    CREATE TABLE IF NOT EXISTS payroll_columns (
      id INT UNSIGNED AUTO_INCREMENT PRIMARY KEY,
      table_name VARCHAR(40) NOT NULL,
      col_index INT NOT NULL,
      col_name VARCHAR(40) NOT NULL,
      display_name VARCHAR(255) NOT NULL
    ) ENGINE=InnoDB;
    """
    execute_one(sql)


def _create_table(table_name: str, col_names: List[str], headers: List[str], no_idx: int):
    cols = []
    for i, c in enumerate(col_names):
        cols.append(f"`{c}` TEXT")
    execute_one(f"DROP TABLE IF EXISTS `{table_name}`;")
    sql = (
        f"CREATE TABLE `{table_name}` ({', '.join(cols)}) "
        "ENGINE=InnoDB ROW_FORMAT=DYNAMIC;"
    )
    execute_one(sql)

    execute_one("DELETE FROM payroll_columns WHERE table_name=%s;", (table_name,))
    for i, (c, h) in enumerate(zip(col_names, headers), start=1):
        execute_one(
            "INSERT INTO payroll_columns (table_name, col_index, col_name, display_name) "
            "VALUES (%s, %s, %s, %s);",
            (table_name, i, c, h or ""),
        )


def _import_file(filepath: str, table_name: str):
    wb = openpyxl.load_workbook(filepath, data_only=True, read_only=True)
    if SHEET_NAME not in wb.sheetnames:
        raise ValueError(f"Sheet '{SHEET_NAME}' not found in {filepath}")
    ws = wb[SHEET_NAME]

    header_row = _find_header_row(ws)
    if header_row == 0:
        raise ValueError(f"Header row with 'NO' not found in {filepath}")

    headers = _build_headers(ws, header_row)
    col_names, no_idx = _sanitize_headers(headers)
    name_idx = -1
    for i, h in enumerate(headers):
        if str(h).strip().lower() == "nome do trabalhador":
            name_idx = i
            break

    data_rows = _read_rows(ws, header_row + 1, len(col_names))

    header_override = None
    if no_idx >= 0:
        for row in data_rows:
            val = "" if row[no_idx] is None else str(row[no_idx]).strip()
            if val.lower() == "no":
                header_override = [("" if v is None else str(v).strip()) for v in row]
                break

    if header_override:
        headers = header_override[:len(headers)]
    headers = [
        ("Retribuicao" if str(h).strip().lower() == "salario base 2024" else h)
        for h in headers
    ]

    _ensure_meta_table()
    _create_table(table_name, col_names, headers, no_idx)

    data_rows = _clean_rows(data_rows, no_idx, name_idx)

    if no_idx >= 0:
        def no_key(row):
            v = row[no_idx]
            try:
                return int(v)
            except Exception:
                return 0
        data_rows.sort(key=no_key)

    placeholders = ", ".join(["%s"] * len(col_names))
    sql = f"INSERT INTO `{table_name}` ({', '.join('`'+c+'`' for c in col_names)}) VALUES ({placeholders})"
    execute_many(sql, data_rows, importer="payroll_chris", filename=os.path.basename(filepath))


def run():
    for month, (fname, table_name) in FILES.items():
        path = os.path.join(PAYROLL_SOURCE_DIR, fname)
        if not os.path.exists(path):
            continue
        _import_file(path, table_name)


if __name__ == "__main__":
    run()

