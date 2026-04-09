"""
importers/import_bim.py
Read BIM Salaries .xlsx files (NIB, Name, Amount format).
"""
import os, glob
import openpyxl
from rich.console import Console
import sys; sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..'))
from config import FOLDERS
from db import get_connection

console = Console()
BIM_DIR = FOLDERS["bim_salaries"]


def _parse_month_from_filename(fname: str) -> int:
    base = os.path.basename(fname)
    try:
        return int(base.split(" ")[0])
    except Exception:
        return 0


def _find_salary_run_id(month: int, year: int = 2026) -> int | None:
    conn = get_connection()
    try:
        cur = conn.cursor(dictionary=True)
        cur.execute("SELECT id FROM salary_runs WHERE month=%s AND year=%s", (month, year))
        row = cur.fetchone()
        return row["id"] if row else None
    finally:
        conn.close()


def import_bim_file(filepath: str):
    month = _parse_month_from_filename(filepath)
    console.print(f"[cyan]BIM file: {os.path.basename(filepath)} (month {month})[/cyan]")

    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb.active

    # Find header row
    header_row = None
    for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if row and any(str(c).strip() == "NIB/Account Number" for c in row if c):
            header_row = i
            break

    if not header_row:
        console.print("[red]  Cannot find NIB header row[/red]")
        return

    salary_run_id = _find_salary_run_id(month)
    loaded = 0

    conn = get_connection()
    try:
        cur = conn.cursor()
        for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
            if not row or row[1] is None:
                continue
            nib = str(row[1]).strip()
            if not nib or not nib.isdigit():
                continue
            name   = str(row[2]).strip() if row[2] else ""
            amount = float(row[3]) if row[3] else 0.0
            desc1  = str(row[4]).strip() if row[4] else ""
            desc2  = str(row[5]).strip() if row[5] else ""
            tx_t   = str(row[6]).strip() if row[6] else "C"
            email  = str(row[7]).strip() if len(row) > 7 and row[7] else None

            cur.execute("""
                INSERT INTO bim_salary_transfers
                  (salary_run_id, nib, employee_name, amount_mzn,
                   description1, description2, transaction_type, email, source_file)
                VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s)
            """, (salary_run_id, nib, name, amount, desc1, desc2, tx_t, email,
                  os.path.basename(filepath)))
            loaded += 1

        conn.commit()
        console.print(f"[green]  ✓ {loaded} BIM transfer lines imported[/green]")
    finally:
        conn.close()


def run():
    console.print("[bold blue]═══ BIM Salary Transfer Importer ═══[/bold blue]")
    files = sorted(glob.glob(os.path.join(BIM_DIR, "*.xlsx")))
    if not files:
        console.print(f"[yellow]No files found in {BIM_DIR}[/yellow]")
        return
    for f in files:
        import_bim_file(f)


if __name__ == "__main__":
    run()
