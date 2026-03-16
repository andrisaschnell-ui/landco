"""
importers/import_shareholders.py
Read Shareholders 2026 .xlsx files (Summary, Income, Expenses sheets).
"""
import os, glob, re
import openpyxl
from datetime import datetime
from rich.console import Console
import sys; sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..'))
from config import FOLDERS
from db import get_connection, fetch_all

console = Console()
SH_DIR = FOLDERS["shareholders"]

SHAREHOLDER_MAP = {
    "LUZ":    1,
    "TAFY":   1,
    "COCO":   2,
    "COHEN":  2,
    "AURORA": 3,
    "STEAD":  3,
    "CAJU":   4,
    "KEVIN":  4,
}

MONTH_MAP = {
    "JANUARY": 1, "JANEIRIO": 1, "JAN": 1,
    "FEBRUARY": 2, "FEB": 2,
    "MARCH": 3, "MAR": 3,
    "APRIL": 4,
    "MAY": 5,
    "JUNE": 6,
    "JULY": 7,
    "AUGUST": 8,
    "SEPTEMBER": 9,
    "OCTOBER": 10,
    "NOVEMBER": 11,
    "DECEMBER": 12,
}


def _detect_shareholder(fname: str) -> int | None:
    fname_up = os.path.basename(fname).upper()
    for key, sh_id in SHAREHOLDER_MAP.items():
        if key in fname_up:
            return sh_id
    return None


def _safe(val) -> float:
    try:
        if val is None or str(val).strip() in ("#REF!", "#N/A", "#VALUE!", ""):
            return 0.0
        return float(str(val).replace(",", "").strip())
    except Exception:
        return 0.0


def import_shareholder_file(filepath: str):
    sh_id = _detect_shareholder(filepath)
    if not sh_id:
        console.print(f"[yellow]  Cannot identify shareholder: {filepath}[/yellow]")
        return

    # Determine month from filename prefix (e.g. "01 COCO COHEN2026.xlsx" = January)
    base = os.path.basename(filepath)
    try:
        month_num = int(base.split(" ")[0])
    except Exception:
        month_num = 0

    year = 2026
    console.print(f"[cyan]Shareholder: {base} → sh_id={sh_id}, month={month_num}[/cyan]")

    wb = openpyxl.load_workbook(filepath, data_only=True, read_only=True)

    # ── SUMMERY sheet ──────────────────────────────────────────
    summary_sheet = None
    for sn in wb.sheetnames:
        if "SUMM" in sn.upper():
            summary_sheet = sn
            break

    if summary_sheet:
        ws = wb[summary_sheet]
        conn = get_connection()
        try:
            cur = conn.cursor()
            opening_bal = 0.0
            monthly_income = {}
            monthly_expenses = {}

            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row or row[0] is None:
                    continue
                label = str(row[0]).strip().upper()

                if "OPENING BALANCE" in label or "OPENING" in label:
                    opening_bal = _safe(row[1])

                month_num_s = MONTH_MAP.get(label)
                if month_num_s:
                    income_val  = _safe(row[1]) if row[1] else 0.0
                    expense_val = _safe(row[4]) if len(row) > 4 and row[4] else 0.0
                    if income_val:
                        monthly_income[month_num_s] = income_val
                    if expense_val:
                        monthly_expenses[month_num_s] = expense_val

            # Upsert shareholder_balances
            for m, inc in monthly_income.items():
                exp = monthly_expenses.get(m, 0.0)
                # Use opening + income - expenses for running balance
                cur.execute("""
                    INSERT INTO shareholder_balances
                      (shareholder_id, month, year,
                       opening_balance_mzn, total_income_mzn,
                       total_expenses_mzn,  closing_balance_mzn)
                    VALUES (%s,%s,%s,%s,%s,%s,%s)
                    ON DUPLICATE KEY UPDATE
                      total_income_mzn   = VALUES(total_income_mzn),
                      total_expenses_mzn = VALUES(total_expenses_mzn),
                      closing_balance_mzn = opening_balance_mzn + VALUES(total_income_mzn) - VALUES(total_expenses_mzn)
                """, (sh_id, m, year, opening_bal, inc, exp,
                      opening_bal + inc - exp))

            conn.commit()
            console.print(f"  [green]✓ Balance summary loaded ({len(monthly_income)} months)[/green]")
        finally:
            conn.close()

    # ── INCOME sheet ──────────────────────────────────────────
    inc_sheet = None
    for sn in wb.sheetnames:
        if "INCOME" in sn.upper():
            inc_sheet = sn
            break

    if inc_sheet:
        ws = wb[inc_sheet]
        conn = get_connection()
        loaded = 0
        try:
            cur = conn.cursor()
            current_month = month_num
            for row in ws.iter_rows(min_row=3, values_only=True):
                if not row:
                    continue
                # Month label in col 0
                if row[0] and str(row[0]).strip().upper() in MONTH_MAP:
                    current_month = MONTH_MAP[str(row[0]).strip().upper()]
                    continue

                if not row[1] and not row[2]:
                    continue

                desc = str(row[1]).strip() if row[1] else ""
                amt_mzn = _safe(row[2])
                amt_usd = _safe(row[3]) if len(row) > 3 else 0.0
                rate    = _safe(row[4]) if len(row) > 4 else None

                if amt_mzn == 0 and amt_usd == 0:
                    continue
                if "OPENING BALANCE" in desc.upper():
                    tx_type = "opening_balance"
                else:
                    tx_type = "income"

                # Get property_id for this shareholder
                props = fetch_all("SELECT id FROM properties WHERE shareholder_id=%s LIMIT 1", (sh_id,))
                prop_id = props[0]["id"] if props else None

                cur.execute("""
                    INSERT INTO shareholder_transactions
                      (shareholder_id, property_id, month, year,
                       description, tx_type, amount_mzn, amount_usd,
                       exchange_rate, source_file)
                    VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
                """, (sh_id, prop_id, current_month, year, desc, tx_type,
                      amt_mzn, amt_usd or None, rate or None,
                      os.path.basename(filepath)))
                loaded += 1

            conn.commit()
            console.print(f"  [green]✓ {loaded} income transactions loaded[/green]")
        finally:
            conn.close()


def run():
    console.print("[bold blue]═══ Shareholder Importer ═══[/bold blue]")
    files = sorted(glob.glob(os.path.join(SH_DIR, "*.xlsx")))
    if not files:
        console.print(f"[yellow]No shareholder files found in {SH_DIR}[/yellow]")
        return
    for f in files:
        if "New Worksheet" in f:
            console.print(f"  [dim]Skipping duplicate: {os.path.basename(f)}[/dim]")
            continue
        import_shareholder_file(f)


if __name__ == "__main__":
    run()

