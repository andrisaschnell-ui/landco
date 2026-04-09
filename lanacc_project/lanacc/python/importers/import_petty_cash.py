"""
importers/import_petty_cash.py
Read Petty Cash .xlsx files (fuel, parks, USD cash sheets).
"""
import os, glob
import openpyxl
from datetime import datetime
from rich.console import Console
import sys; sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..'))
from config import FOLDERS
from db import get_connection

console = Console()
ACCOUNTS_DIR = FOLDERS["accounts"]

SHEET_ACCOUNT_MAP = {
    "RSDFUEL": "RSDFuel",
    "RSD":     "RSDFuel",
    "FUEL":    "RSDFuel",
    "PARKS":   "PARKS_FEES",
    "USD":     "USD_CASH",
    "PRE-PAID":"PRE_PAID",
    "PETTY":   "PETTY_CASH",
    "MTN":     "MTN_CASH",
}


def _safe(val) -> float:
    try:
        if val is None or str(val).strip() in ("#REF!", "#N/A", "#VALUE!", ""):
            return 0.0
        return float(str(val).replace(",", "").strip())
    except Exception:
        return 0.0


def _parse_month_from_filename(fname: str) -> int:
    base = os.path.basename(fname)
    try:
        return int(base.split(" ")[0])
    except Exception:
        return 0


def import_petty_cash_file(filepath: str):
    month = _parse_month_from_filename(filepath)
    year = 2026
    console.print(f"[cyan]Petty cash: {os.path.basename(filepath)} (month {month})[/cyan]")

    wb = openpyxl.load_workbook(filepath, data_only=True)
    conn = get_connection()
    try:
        cur = conn.cursor()
        total_loaded = 0

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            sheet_up = sheet_name.strip().upper().replace(" ", "")

            # Detect account type from sheet name
            account_name = None
            for key, val in SHEET_ACCOUNT_MAP.items():
                if key in sheet_up:
                    account_name = val
                    break
            if not account_name:
                account_name = sheet_name.strip()

            # Determine currency
            currency = "USD" if "USD" in sheet_up or "PARKS" in sheet_up else "MZN"

            # Find header row with DATE column
            header_row = None
            for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
                if row and any(str(c).strip().upper() == "DATE" for c in row if c):
                    header_row = i
                    break
            if not header_row:
                continue

            for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
                if not row or row[0] is None:
                    continue
                # Row[0] should be a date
                tx_date = None
                if isinstance(row[0], datetime):
                    tx_date = row[0].date()
                else:
                    try:
                        tx_date = datetime.strptime(str(row[0]).strip(), "%Y-%m-%d").date()
                    except Exception:
                        continue

                desc  = str(row[1]).strip() if row[1] else ""
                cr    = _safe(row[2])
                dr    = _safe(row[3])
                bal   = _safe(row[4]) if len(row) > 4 else 0.0

                if cr == 0 and dr == 0:
                    continue

                cur.execute("""
                    INSERT INTO petty_cash_transactions
                      (account_name, tx_date, month, year,
                       description, credit, debit, balance,
                       currency, source_file)
                    VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
                """, (account_name, tx_date, month, year,
                      desc, cr, dr, bal, currency,
                      os.path.basename(filepath)))
                total_loaded += 1

        conn.commit()
        console.print(f"  [green]✓ {total_loaded} petty cash rows loaded[/green]")
    finally:
        conn.close()


def run():
    console.print("[bold blue]═══ Petty Cash Importer ═══[/bold blue]")
    pattern = os.path.join(ACCOUNTS_DIR, "**", "*PETTY CASH*.xlsx")
    files = sorted(glob.glob(pattern, recursive=True))
    if not files:
        console.print(f"[yellow]No petty cash files found[/yellow]")
        return
    for f in files:
        import_petty_cash_file(f)


if __name__ == "__main__":
    run()
