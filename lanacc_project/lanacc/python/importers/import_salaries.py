"""
importers/import_salaries.py
Read Landco Salaries 2026 / Folha de Salarios .xlsx files.
"""
import os, glob
import openpyxl
from datetime import datetime
from rich.console import Console
import sys; sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..'))
from config import FOLDERS
from db import get_connection, execute_many, execute_one, fetch_all

console = Console()

SALARY_DIR = os.path.join(FOLDERS["salaries"], "Landco Salaries 2026")

# Column positions in 'Folha de salarios' sheet (0-indexed after header row)
# Based on actual file analysis
COL_NO        = 0
COL_INIT      = 1
COL_NAME      = 2
COL_BASE      = 7   # Salario Base
COL_FOOD      = 9   # Alimentacao
COL_BACK_PAY  = 10
COL_DAYS      = 11
COL_MONTHLY   = 12
COL_NIGHTSH   = 13
COL_GROSS     = 15  # approximate — varies per sheet


def _parse_month_year_from_filename(fname: str):
    """Extract month number from filenames like '01 Salary sheet for Landco.xlsx'"""
    base = os.path.basename(fname)
    try:
        month_num = int(base.split(" ")[0])
        year = 2026
        return month_num, year
    except Exception:
        return None, None


def _get_or_create_salary_run(month: int, year: int, source_file: str) -> int:
    conn = get_connection()
    try:
        cur = conn.cursor(dictionary=True)
        cur.execute("SELECT id FROM salary_runs WHERE month=%s AND year=%s", (month, year))
        row = cur.fetchone()
        if row:
            return row["id"]
        cur.execute("""
            INSERT INTO salary_runs (month, year, company, nuit, source_file, status)
            VALUES (%s, %s, 'LANDCO LIMITADA', '400122792', %s, 'draft')
        """, (month, year, source_file))
        conn.commit()
        return cur.lastrowid
    finally:
        conn.close()


def _get_employee_id(name: str) -> int | None:
    rows = fetch_all("SELECT id FROM employees WHERE name = %s", (name.strip(),))
    if rows:
        return rows[0]["id"]
    # Insert unknown employee
    conn = get_connection()
    try:
        cur = conn.cursor()
        cur.execute("INSERT INTO employees (name, active) VALUES (%s, 1)", (name.strip(),))
        conn.commit()
        return cur.lastrowid
    finally:
        conn.close()


def _safe_decimal(val) -> float:
    try:
        if val is None or val == "" or str(val).strip() in ("#REF!", "#N/A", "#VALUE!"):
            return 0.0
        return float(str(val).replace(",", "").strip())
    except Exception:
        return 0.0


def import_salary_file(filepath: str):
    month, year = _parse_month_year_from_filename(filepath)
    if not month:
        console.print(f"[yellow]  Cannot determine month from filename: {filepath}[/yellow]")
        return

    console.print(f"[cyan]Processing salary file: {os.path.basename(filepath)} ({month}/{year})[/cyan]")

    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb["Folha de salarios"]

    salary_run_id = _get_or_create_salary_run(month, year, os.path.basename(filepath))

    # Find header row (row with 'NOME DO TRABALHADOR')
    header_row = None
    for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if row and any(str(c).strip() == "NOME DO TRABALHADOR" for c in row if c):
            header_row = i
            break

    if not header_row:
        console.print("[red]  Could not find header row — skipping[/red]")
        return

    lines_added = 0
    conn = get_connection()
    try:
        cur = conn.cursor()
        for row in ws.iter_rows(min_row=header_row + 2, values_only=True):
            if not row or row[COL_NO] is None:
                continue
            try:
                int(row[COL_NO])
            except (ValueError, TypeError):
                continue

            name = str(row[COL_NAME]).strip() if row[COL_NAME] else None
            if not name or name.lower() in ("none", ""):
                continue

            emp_id = _get_employee_id(name)
            base = _safe_decimal(row[COL_BASE])
            food = _safe_decimal(row[COL_FOOD])
            back_pay = _safe_decimal(row[COL_BACK_PAY])
            days = row[COL_DAYS] if row[COL_DAYS] else 30
            gross = base + food + back_pay

            # Mozambique statutory deductions
            inss_emp = round(gross * 0.03, 2)
            inss_emp_r = round(gross * 0.04, 2)
            sindicate = round(gross * 0.01, 2)
            taxable = gross - inss_emp
            irps = _calc_irps(taxable)
            net = gross - inss_emp - sindicate - irps

            cur.execute("""
                INSERT INTO salary_lines
                  (salary_run_id, employee_id, base_salary, food_allowance,
                   back_pay, gross_salary, inss_employee, inss_employer,
                   sindicate, irps, net_salary, days_worked, source_file)
                VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
                ON DUPLICATE KEY UPDATE gross_salary=VALUES(gross_salary)
            """, (salary_run_id, emp_id, base, food, back_pay, gross,
                  inss_emp, inss_emp_r, sindicate, irps, net, days,
                  os.path.basename(filepath)))
            lines_added += 1

        # Update salary run totals
        cur.execute("""
            UPDATE salary_runs sr
            JOIN (
              SELECT salary_run_id,
                     SUM(gross_salary)   gross,
                     SUM(net_salary)     net,
                     SUM(inss_employee)  ie,
                     SUM(inss_employer)  ir,
                     SUM(irps)           ix
              FROM salary_lines
              WHERE salary_run_id = %s
              GROUP BY salary_run_id
            ) t ON t.salary_run_id = sr.id
            SET sr.total_gross_mzn     = t.gross,
                sr.total_net_mzn       = t.net,
                sr.total_inss_employee = t.ie,
                sr.total_inss_employer = t.ir,
                sr.total_irps          = t.ix
            WHERE sr.id = %s
        """, (salary_run_id, salary_run_id))

        conn.commit()
        console.print(f"[green]  ✓ {lines_added} salary lines imported[/green]")
    finally:
        conn.close()


def _calc_irps(taxable: float) -> float:
    """Mozambique IRPS brackets 2025/2026 (approximate)."""
    if taxable <= 20249:
        return 0.0
    elif taxable <= 30374:
        return round((taxable - 20249) * 0.10, 2)
    elif taxable <= 40499:
        return round(1012.50 + (taxable - 30374) * 0.15, 2)
    elif taxable <= 60749:
        return round(2531.25 + (taxable - 40499) * 0.20, 2)
    elif taxable <= 101249:
        return round(6581.25 + (taxable - 60749) * 0.25, 2)
    else:
        return round(16706.25 + (taxable - 101249) * 0.32, 2)


def run():
    console.print("[bold blue]═══ Salary Importer ═══[/bold blue]")
    files = sorted(glob.glob(os.path.join(SALARY_DIR, "*.xlsx")))
    if not files:
        console.print(f"[yellow]No files found in {SALARY_DIR}[/yellow]")
        return
    for f in files:
        import_salary_file(f)


if __name__ == "__main__":
    run()
